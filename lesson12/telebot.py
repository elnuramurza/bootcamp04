# from random import randint
from openpyxl import load_workbook

# a = randint(2, 10)
# print(a)

excel_file = load_workbook("students1.xlsx")
page = excel_file["Лист1"]

end_row = len(page["A"]) # 7
for i in range(2, end_row + 1): # 8
    mark_1 = page[f"E{i}"].value # page["E2"]
    mark_2 = page[f"F{i}"].value
    mark_3 = page[f"G{i}"].value
    mediana = (mark_1 + mark_2 + mark_3) / 3
    page[f"H{i}"] = mediana
excel_file.save("students1.xlsx") 
 
import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime

class TelegramBot:
    # создание бота
    def __init__(self, token):
        self.token = token
        self.url = f"https://api.telegram.org/bot{token}/"

    # получение обновлений (все последние сообщения)
    def get_updates(self):
        response = requests.get(f"{self.url}getUpdates")
        data = response.json()
        return data

    # получить текст последнего сообщения
    def get_last(self):
        data = self.get_updates()
        return data["result"][-1]["message"]["text"]

    # отправка сообщения
    def send_message(self, chat_id, text):
        requests.get(f"{self.url}sendMessage?chat_id={chat_id}&text={text}")

    # адресная справка
    def sync_chats(self):
        data = self.get_updates()
        chats = {}
        for update in data["result"]:
            key = update["message"]["chat"]["id"]
            value = update["message"]["from"]["first_name"]
            chats[key] = value
        self.chats = chats
        return chats

    def spam(self, text):
        for chat, name in self.chats.items():
            text_message = f"Hello, {name}! {text}"
            self.send_message(chat, text_message)

    # создания excel с чатами
    def create_chats_excel(self, file_name=None):
        if not file_name:
            # генерируем название
            now = int(datetime.now().timestamp())
            file_name = f"chats{now}.xlsx"

        self.file_name = file_name
        # создаём новый excel файл
        new_excel = Workbook()
        page = new_excel.active

        # записываем чаты в excel
        chats = self.sync_chats()
        for key, value in chats.items():
            page.append([key, value])

        # сохраняем
        new_excel.save(file_name)
        return file_name

    # чтение записей с excel
    def get_chats_excel(self):
        excel_file = load_workbook(self.file_name)
        page = excel_file[excel_file.sheetnames[0]]
        chats = {}
        for row in page:
            chat_id = row[0].value
            name = row[1].value
            chats[chat_id] = name
        return chats

    def send_photo(self, chat_id, photo_url):
        requests.get(f"{self.url}sendPhoto?chat_id={chat_id}&photo={photo_url}")
 

from bot_class import TelegramBot

token = "6450383163:AAHb6Do-tMkORRBWdQjAyZrXyWyq1-1teC0"
bot = TelegramBot(token)

chat_id = "270623373" # Kaium chat id
# bot.send_message(chat_id, "hello from bot oop")

# bot.sync_chats()
# bot.spam("test Kaium")

# file_name = "test_1.xlsx"
# bot.create_chats_excel(file_name)
# print(bot.get_chats_excel())

photo_url = "https://leonardo.osnova.io/5569a907-bbcf-5657-9fcb-a472aa78165f/-/preview/2100/-/format/webp/"
bot.send_photo(chat_id, photo_url)
 